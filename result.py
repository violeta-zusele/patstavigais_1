from openpyxl import Workbook, load_workbook 
wb=load_workbook('tests/test1.xlsx')
ws=wb.active
total=0
#write your code here
max_row = ws.max_row
for row in range(2,max_row+1):  
    try:
        rate = float(ws['B'+str(row)].value)
        hours = float(ws['C'+str(row)].value)
        if(type(rate)!=str and type(hours)!=str):
            salary = rate * hours
            ws['D'+str(row)].value = salary
            if salary > 3000:
                total +=1
    except (ValueError, TypeError):
        continue
print(total)
wb.close()